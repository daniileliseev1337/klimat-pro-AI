"""Excel-дашборд (openpyxl)."""

from __future__ import annotations

from pathlib import Path
from typing import Iterable, Optional

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from .db import Database
from .models import Listing, PricePoint, Score


_HEADER_FONT = Font(bold=True, color="FFFFFF")
_HEADER_FILL = PatternFill("solid", fgColor="4472C4")
_SCORE_HIGH_FILL = PatternFill("solid", fgColor="C6EFCE")
_SCORE_LOW_FILL = PatternFill("solid", fgColor="FFC7CE")


def _set_header(ws, row: int, headers: Iterable[str]) -> None:
    for col, name in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col, value=name)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")


def _autosize(ws) -> None:
    for col in ws.columns:
        max_len = 0
        column_letter = get_column_letter(col[0].column)
        for cell in col:
            v = cell.value
            if v is None:
                continue
            max_len = max(max_len, min(50, len(str(v))))
        ws.column_dimensions[column_letter].width = max(10, max_len + 2)


def export(db: Database, out_path: str | Path, *, score_threshold_high: float = 70.0) -> Path:
    out = Path(out_path)
    out.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()

    listings = db.all_listings()
    scores = {s.listing_id: s for s in db.all_scores()}
    latest_prices = {l.id: db.latest_price(l.id) for l in listings}

    _write_listings_sheet(wb.active, listings, latest_prices, scores, score_threshold_high)
    _write_scoring_sheet(wb.create_sheet("Скоринг"), listings, scores, latest_prices, score_threshold_high)
    _write_history_sheet(wb.create_sheet("История цен"), db, listings)
    _write_comparison_sheet(wb.create_sheet("Сравнение"))

    wb.save(out)
    return out


def _write_listings_sheet(
    ws,
    listings: list[Listing],
    latest_prices: dict[str, Optional[PricePoint]],
    scores: dict[str, Score],
    threshold_high: float,
) -> None:
    ws.title = "Лоты"
    headers = [
        "ID", "Источник", "Статус", "URL", "Заголовок", "Адрес", "Город", "Район",
        "Комнат", "Площадь, м²", "Этаж", "Этажей", "Год", "Тип дома", "Ремонт",
        "Метро", "До метро, мин", "Продавец", "Фото",
        "Цена, ₽", "₽/м²", "Скоринг",
        "Первое появление", "Последнее появление", "Заметки",
    ]
    _set_header(ws, 1, headers)
    for i, l in enumerate(listings, start=2):
        price_point = latest_prices.get(l.id)
        price = price_point.price if price_point else None
        ppm = round(price / l.area_total, 0) if price and l.area_total else None
        score_val = scores[l.id].score if l.id in scores else None
        row = [
            l.id, l.source, l.status, l.url, l.title, l.address, l.city, l.district,
            l.rooms, l.area_total, l.floor, l.floors_total, l.year_built,
            l.building_type, l.renovation, l.metro_name, l.metro_distance_min,
            l.seller_type, l.photos_count,
            price, ppm, score_val,
            l.first_seen.isoformat(timespec="seconds"),
            l.last_seen.isoformat(timespec="seconds"),
            l.notes,
        ]
        for col, val in enumerate(row, start=1):
            cell = ws.cell(row=i, column=col, value=val)
            if col == 22 and isinstance(val, (int, float)):
                if val >= threshold_high:
                    cell.fill = _SCORE_HIGH_FILL
                elif val < 40:
                    cell.fill = _SCORE_LOW_FILL
    ws.freeze_panes = "B2"
    _autosize(ws)


def _write_scoring_sheet(
    ws,
    listings: list[Listing],
    scores: dict[str, Score],
    latest_prices: dict[str, Optional[PricePoint]],
    threshold_high: float,
) -> None:
    if not scores:
        ws["A1"] = "Скоринг ещё не считался. Выполни: apartment-tracker score"
        return

    sample = next(iter(scores.values()))
    crit_keys = list(sample.breakdown.keys())
    headers = ["ID", "Заголовок", "Адрес", "Цена, ₽", "Скоринг"] + crit_keys
    _set_header(ws, 1, headers)

    rows = []
    for l in listings:
        sc = scores.get(l.id)
        if not sc:
            continue
        price_point = latest_prices.get(l.id)
        price = price_point.price if price_point else None
        row = [l.id, l.title, l.address, price, sc.score]
        for k in crit_keys:
            row.append(sc.breakdown.get(k))
        rows.append(row)

    rows.sort(key=lambda r: r[4] or 0, reverse=True)
    for i, r in enumerate(rows, start=2):
        for col, val in enumerate(r, start=1):
            cell = ws.cell(row=i, column=col, value=val)
            if col == 5 and isinstance(val, (int, float)):
                if val >= threshold_high:
                    cell.fill = _SCORE_HIGH_FILL
                elif val < 40:
                    cell.fill = _SCORE_LOW_FILL
    ws.freeze_panes = "B2"
    _autosize(ws)


def _write_history_sheet(ws, db: Database, listings: list[Listing]) -> None:
    headers = ["ID", "Заголовок", "Время", "Цена, ₽", "Δ к предыдущей"]
    _set_header(ws, 1, headers)
    row = 2
    for l in listings:
        history = db.price_history(l.id)
        if not history:
            continue
        prev = None
        for p in history:
            delta = p.price - prev if prev is not None else None
            ws.cell(row=row, column=1, value=l.id)
            ws.cell(row=row, column=2, value=l.title)
            ws.cell(row=row, column=3, value=p.seen_at.isoformat(timespec="seconds"))
            ws.cell(row=row, column=4, value=p.price)
            ws.cell(row=row, column=5, value=delta)
            row += 1
            prev = p.price
    ws.freeze_panes = "A2"
    _autosize(ws)


def _write_comparison_sheet(ws) -> None:
    """Пустой лист-шаблон для side-by-side сравнения 3-5 лотов вручную."""
    ws["A1"] = "Лист для ручного сравнения: вставь ID лотов из листа «Лоты» в B1:F1 и формулы VLOOKUP подтянут поля."
    ws["A1"].font = Font(italic=True, color="666666")

    headers = ["Параметр", "Лот 1", "Лот 2", "Лот 3", "Лот 4", "Лот 5"]
    _set_header(ws, 3, headers)

    params = [
        "ID", "URL", "Адрес", "Комнат", "Площадь, м²", "Этаж/Этажей", "Год",
        "Ремонт", "Метро / минут", "Цена, ₽", "₽/м²", "Скоринг", "Мои плюсы", "Мои минусы",
    ]
    for i, p in enumerate(params, start=4):
        cell = ws.cell(row=i, column=1, value=p)
        cell.font = Font(bold=True)
    ws.column_dimensions["A"].width = 22
    for col in range(2, 7):
        ws.column_dimensions[get_column_letter(col)].width = 30
